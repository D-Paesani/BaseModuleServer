from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image
from openpyxl.styles.colors import Color
from openpyxl.styles.fills import PatternFill
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from os import remove
from flask import send_file

red = Color(rgb='dc3545')
green = Color(rgb='198754')
gray = Color(rgb='6c757d')
warn = Color(rgb='fffb04')
bg_red = PatternFill(patternType='solid', fgColor=red) #colore cella background
bg_green = PatternFill(patternType='solid', fgColor=green) #colore cella background
bg_gray = PatternFill(patternType='solid', fgColor=gray) #colore cella background
bg_warn = PatternFill(patternType='solid', fgColor=warn) #colore cella background

txt_foreground = Font(color='ffffff', bold=True, size=12) #font color white bold
txt_foreground_black = Font(color='000000', bold=True, size=12) #font color black bold
txt_pry = Font(color='0d6efd', bold=True, size=14) #testo head color primary
txt_sw = Font(color='4154f1', bold=True, size=13) #per il contenuto testo speciale es. SW1,SW2 module id
txt_sub = Font(bold=True, size=12) #sotto head color black e bold
txt_cont = Font(size=12) #per il contenuto tabella
txt_data = Font(size=9, bold=True, color='008080') #per la data
txt_sub_red = Font(color='dc3545', bold=True, size=12)
txt_sub_green = Font(color='198754', bold=True, size=12)
txt_sub_gray = Font(color='6c757d', bold=True, size=12)
align_centr = Alignment(horizontal='center', vertical='center')

inline_red = InlineFont(txt_sub_red)
inline_green = InlineFont(txt_sub_green)
inline_gray = InlineFont(txt_sub_gray)